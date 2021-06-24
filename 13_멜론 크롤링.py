import urllib.request as req
from bs4 import BeautifulSoup
import os
import openpyxl
import datetime
from openpyxl.drawing.image import Image
import re # 추가

# 이미지 저장할 폴더 생성
if not os.path.exists("./멜론이미지"):
    os.mkdir("./멜론이미지")

header = req.Request("https://www.melon.com/chart/index.htm", headers={"User-Agent":"Mozilla/5.0"})
code = req.urlopen(header)
soup = BeautifulSoup(code, "html.parser")
title = soup.select("div.ellipsis.rank01 > span > a")
name = soup.select("div.ellipsis.rank02 > span")
album = soup.select("div.ellipsis.rank03 > a")
img = soup.select("a.image_typeAll > img")

# 엑셀 파일 생성
if not os.path.exists("./멜론_크롤링.xlsx"):
    openpyxl.Workbook().save("./멜론_크롤링.xlsx")

book = openpyxl.load_workbook("./멜론_크롤링.xlsx")
# 쓸데 없는 시트는 삭제하기
if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])
sheet = book.create_sheet()
now = datetime.datetime.now()
sheet.title = "{}년 {}월 {}일 {}시 {}분 {}초".format(now.year, now.month, now.day, now.hour, now.minute, now.second)
row_num = 1
# 열 너비 조절
sheet.column_dimensions["A"].width = 15
sheet.column_dimensions["B"].width = 50
sheet.column_dimensions["C"].width = 29
sheet.column_dimensions["D"].width = 47
for i in range(len(title)):
    img_file_name = "./멜론이미지/{}.png".format(re.sub("[\\\/:*?\"<>\|]", " ", album[i].string))  # 수정
    req.urlretrieve(img[i].attrs["src"], img_file_name) # 수정
    # 엑셀에 크롤링 결과 출력
    img_for_excel = Image(img_file_name)
    sheet.add_image(img_for_excel, "A{}".format(row_num))
    sheet.cell(row=row_num, column=2).value = title[i].string
    sheet.cell(row=row_num, column=3).value = name[i].text
    sheet.cell(row=row_num, column=4).value = album[i].string
    sheet.row_dimensions[row_num].height = 90 # 행높이 조절.
    book.save("./멜론_크롤링.xlsx")
    print("{}위. {} - {}".format(row_num, title[i].string, name[i].text))
    row_num+=1
